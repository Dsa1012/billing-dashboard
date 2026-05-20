"""
excel_reader.py — Nucleo de lectura del Excel.
Lee color de FUENTE: verde=#00B050 (facturado), azul=#0000CC (proyeccion).
"""

from __future__ import annotations
import logging
import shutil
import tempfile
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Optional
import openpyxl

from app.config import (
    EXCEL_FOLDER, EXCEL_EXTENSIONS, DATA_START_ROW,
    COL_CLIENTE, COL_DESCRIPCION, COL_CC, COL_GERENTE, COL_COTIZACION,
    YEAR_START_COL, YEAR_END_COL, MONTH_NAMES,
    GREEN_FONT_COLORS, BLUE_FONT_COLORS, COLOR_TOLERANCE,
    SUPABASE_URL, SUPABASE_KEY, SUPABASE_BUCKET, SUPABASE_FILE, CLOUD_MODE,
)
from app.models import (
    MonthEntry, RowRecord, ClienteSummary, GerenteSummary,
    CCSummary, MonthSummary, KPIs, DashboardData,
)

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────
# MODO NUBE — descarga desde Supabase Storage
# ──────────────────────────────────────────────────────────────

def download_from_supabase() -> Optional[Path]:
    """
    Descarga el Excel desde Supabase Storage y lo guarda en un archivo temporal.
    Retorna la ruta al temporal, o None si falla.
    """
    if not CLOUD_MODE:
        return None
    url = f"{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/{SUPABASE_FILE}"
    logger.info("Descargando Excel desde Supabase: %s", url)
    try:
        req = urllib.request.Request(url, headers={"apikey": SUPABASE_KEY})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = resp.read()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(data)
            tmp_path = Path(tmp.name)
        logger.info("Descarga OK — %d bytes", len(data))
        return tmp_path
    except Exception as exc:
        logger.error("Error descargando desde Supabase: %s", exc)
        return None


def get_supabase_etag() -> Optional[str]:
    """Obtiene el ETag del archivo en Supabase para detectar cambios sin descargarlo."""
    if not CLOUD_MODE:
        return None
    url = f"{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/{SUPABASE_FILE}"
    try:
        req = urllib.request.Request(url, method="HEAD", headers={"apikey": SUPABASE_KEY})
        with urllib.request.urlopen(req, timeout=10) as resp:
            return resp.headers.get("ETag") or resp.headers.get("Last-Modified")
    except Exception:
        return None


def _hex_to_rgb(h: str):
    h = h.strip().lstrip("#")
    if len(h) == 8: h = h[2:]
    if len(h) != 6: return (0, 0, 0)
    return (int(h[0:2],16), int(h[2:4],16), int(h[4:6],16))


def _dist(c1, c2):
    return max(abs(c1[0]-c2[0]), abs(c1[1]-c2[1]), abs(c1[2]-c2[2]))


def _classify_font_color(hex_color: Optional[str]) -> str:
    if not hex_color:
        return "sin_color"
    h = hex_color.strip().lstrip("#")
    if len(h) == 8: h = h[2:]
    if len(h) != 6: return "sin_color"
    if h.upper() in ("000000","FFFFFF"): return "sin_color"
    rgb = _hex_to_rgb(h)
    for ref in GREEN_FONT_COLORS:
        if _dist(rgb, _hex_to_rgb(ref)) <= COLOR_TOLERANCE:
            return "facturado"
    for ref in BLUE_FONT_COLORS:
        if _dist(rgb, _hex_to_rgb(ref)) <= COLOR_TOLERANCE:
            return "proyeccion"
    r, g, b = rgb
    if g > r + 40 and g > b + 40 and g > 100: return "facturado"
    if b > r + 40 and b > g + 40 and b > 100: return "proyeccion"
    return "sin_color"


def _get_cell_font_color_hex(cell) -> Optional[str]:
    try:
        font = cell.font
        if font is None: return None
        color = font.color
        if color is None: return None
        if color.type == "rgb": return color.rgb
        return None
    except Exception:
        return None


def find_latest_excel(folder: str = EXCEL_FOLDER) -> Optional[Path]:
    p = Path(folder)
    if not p.exists():
        logger.warning("Carpeta no encontrada: %s", folder)
        return None
    candidates = [
        f for f in p.iterdir()
        if f.is_file() and f.suffix.lower() in EXCEL_EXTENSIONS
        and not f.name.startswith("~$")
    ]
    if not candidates:
        logger.warning("Sin archivos Excel en: %s", folder)
        return None
    latest = max(candidates, key=lambda f: f.stat().st_mtime)
    logger.info("Archivo: %s (%s)", latest.name,
                datetime.fromtimestamp(latest.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"))
    return latest


def read_excel_data(file_path: Path) -> DashboardData:
    logger.info("Leyendo: %s", file_path)
    # Copia a un temporal para poder leer incluso si Excel tiene el archivo abierto
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)
        shutil.copy2(str(file_path), str(tmp_path))
        wb = openpyxl.load_workbook(str(tmp_path), data_only=True, read_only=False)
    except Exception:
        # Si la copia falla, intenta leer el original directamente
        logger.warning("No se pudo copiar a temporal, leyendo directo.")
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=False)
        tmp_path = None
    ws = wb.active

    oc_start  = YEAR_START_COL + 1
    oc_end    = YEAR_END_COL + 1
    n_months  = oc_end - oc_start + 1
    ex_start  = DATA_START_ROW + 1

    rows_data = []
    for row_idx in range(ex_start, ws.max_row + 1):
        def cv(col0):
            return ws.cell(row=row_idx, column=col0+1).value

        cliente     = str(cv(COL_CLIENTE)     or "").strip()
        descripcion = str(cv(COL_DESCRIPCION) or "").strip()
        cc          = str(cv(COL_CC)          or "").strip()
        gerente     = str(cv(COL_GERENTE)     or "").strip()
        if not cliente and not descripcion and not gerente:
            continue
        try:
            cotizacion = float(cv(COL_COTIZACION) or 0)
        except (ValueError, TypeError):
            cotizacion = 0.0

        meses = []
        total_f = total_p = 0.0
        for m_idx in range(n_months):
            cell = ws.cell(row=row_idx, column=oc_start + m_idx)
            try:
                amount = float(cell.value or 0)
            except (ValueError, TypeError):
                amount = 0.0
            tipo = _classify_font_color(_get_cell_font_color_hex(cell))
            month_name = MONTH_NAMES[m_idx] if m_idx < len(MONTH_NAMES) else f"Mes {m_idx+1}"
            meses.append(MonthEntry(month=month_name, month_index=m_idx, amount=amount, tipo=tipo))
            if tipo == "facturado":  total_f += amount
            elif tipo == "proyeccion": total_p += amount

        dif = cotizacion - total_f - total_p
        pct = (total_f / cotizacion * 100) if cotizacion > 0 else 0.0
        rows_data.append(RowRecord(
            row_number=row_idx,
            cliente=cliente or "Sin cliente",
            descripcion=descripcion,
            cc=cc or "Sin CC",
            gerente=gerente or "Sin gerente",
            cotizacion=cotizacion,
            meses=meses,
            total_facturado=round(total_f, 2),
            total_proyeccion=round(total_p, 2),
            diferencia=round(dif, 2),
            pct_cumplimiento=round(pct, 2),
        ))

    wb.close()
    # Eliminar copia temporal
    if tmp_path and tmp_path.exists():
        try:
            tmp_path.unlink()
        except Exception:
            pass
    if not rows_data:
        logger.warning("Sin datos en el archivo.")
    d = _build_dashboard(rows_data, file_path)
    logger.info("OK: %d registros — Fact: %.0f  Proy: %.0f",
                len(rows_data), d.kpis.total_facturado, d.kpis.total_proyeccion)
    return d


def _build_dashboard(rows, file_path):
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Total a facturar 2026 = suma de TODOS los montos mensuales (BJ:BU),
    # independiente del color de la celda.
    total_anual = sum(m.amount for r in rows for m in r.meses)
    total_fact  = sum(r.total_facturado  for r in rows)
    total_proy  = sum(r.total_proyeccion for r in rows)
    sin_asignar = total_anual - total_fact - total_proy
    pct         = (total_fact / total_anual * 100) if total_anual > 0 else 0.0

    # Mantenemos total_cotizacion para compatibilidad pero no se muestra
    total_cot = sum(r.cotizacion for r in rows)

    kpis = KPIs(
        total_a_facturar=round(total_anual, 2),
        total_facturado=round(total_fact, 2),
        total_proyeccion=round(total_proy, 2),
        sin_asignar=round(sin_asignar, 2),
        pct_cumplimiento=round(pct, 2),
        num_clientes=len({r.cliente for r in rows}),
        num_gerentes=len({r.gerente for r in rows}),
        num_registros=len(rows),
        source_file=file_path.name,
        last_updated=now_str,
        total_cotizacion=round(total_cot, 2),
        diferencia=round(sin_asignar, 2),
    )

    def agg(key_fn):
        d = {}
        for r in rows:
            k = key_fn(r)
            if k not in d:
                d[k] = dict(cotizacion=0.0, facturado=0.0, proyeccion=0.0, extra=set())
            d[k]["cotizacion"]  += r.cotizacion
            d[k]["facturado"]   += r.total_facturado
            d[k]["proyeccion"]  += r.total_proyeccion
        return d

    # Por cliente
    by_c = agg(lambda r: r.cliente)
    for r in rows:
        by_c[r.cliente]["extra"].add(r.gerente)
    by_cliente = []
    for cli, d in sorted(by_c.items(), key=lambda x: -x[1]["facturado"]):
        cot = d["cotizacion"]; fac = d["facturado"]; proy = d["proyeccion"]
        pct2 = (fac / cot * 100) if cot > 0 else 0.0
        by_cliente.append(ClienteSummary(
            cliente=cli, cotizacion=round(cot,2), total_facturado=round(fac,2),
            total_proyeccion=round(proy,2), diferencia=round(cot-fac-proy,2),
            pct_cumplimiento=round(pct2,2), gerentes=sorted(d["extra"]),
        ))

    # Por gerente
    by_g = agg(lambda r: r.gerente)
    for r in rows:
        by_g[r.gerente]["extra"].add(r.cliente)
    by_gerente = []
    for ger, d in sorted(by_g.items(), key=lambda x: -x[1]["facturado"]):
        cot = d["cotizacion"]; fac = d["facturado"]; proy = d["proyeccion"]
        pct2 = (fac / cot * 100) if cot > 0 else 0.0
        by_gerente.append(GerenteSummary(
            gerente=ger, cotizacion=round(cot,2), total_facturado=round(fac,2),
            total_proyeccion=round(proy,2), diferencia=round(cot-fac-proy,2),
            pct_cumplimiento=round(pct2,2), num_clientes=len(d["extra"]),
        ))

    # Por CC
    by_cc_dict = agg(lambda r: r.cc)
    by_cc = []
    for cc_name, d in sorted(by_cc_dict.items(), key=lambda x: -x[1]["facturado"]):
        cot = d["cotizacion"]; fac = d["facturado"]; proy = d["proyeccion"]
        pct2 = (fac / cot * 100) if cot > 0 else 0.0
        by_cc.append(CCSummary(
            cc=cc_name, cotizacion=round(cot,2), total_facturado=round(fac,2),
            total_proyeccion=round(proy,2), diferencia=round(cot-fac-proy,2),
            pct_cumplimiento=round(pct2,2),
        ))

    # Por mes
    mf = [0.0]*12; mp = [0.0]*12
    for r in rows:
        for m in r.meses:
            if 0 <= m.month_index < 12:
                if m.tipo == "facturado":   mf[m.month_index] += m.amount
                elif m.tipo == "proyeccion": mp[m.month_index] += m.amount

    by_month = [
        MonthSummary(month=MONTH_NAMES[i], month_index=i,
                     facturado=round(mf[i],2), proyeccion=round(mp[i],2),
                     total=round(mf[i]+mp[i],2))
        for i in range(12)
    ]

    return DashboardData(
        kpis=kpis, rows=rows,
        by_cliente=by_cliente, by_gerente=by_gerente,
        by_cc=by_cc, by_month=by_month,
    )
